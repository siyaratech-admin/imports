#!/usr/bin/env python3
"""
Create a corrected import sheet for variant items that failed in Item(12).csv.

Strategy:
1. Read Fixed_Variants.xlsx → lookup table by Item Code (correct attributes from our parser)
2. Read Item(12).csv → find all variant rows (Variant Of not empty)
3. For each variant item:
   a. If found in Fixed_Variants → use those rows (authoritative correct attributes)
   b. If NOT found → use Item(12).csv rows as-is (they look reasonable)
4. Write to Fixed_Item12_Variants.xlsx

Usage:
    python3 "Material Library/create_item12_corrected.py"
"""
import os, csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

FIXED_VARIANTS_PATH = os.path.join(BASE_DIR, "Fixed_Variants.xlsx")
ITEM12_PATH         = os.path.join(BASE_DIR, "Item(12).csv")
OUTPUT_PATH         = os.path.join(BASE_DIR, "Fixed_Item12_Variants.xlsx")

# ── Column headers (matching ERPNext import format) ───────────────────────────
HEADERS = [
    "Item Code", "Item Group", "Default Unit of Measure", "Has Variants",
    "Maintain Stock", "Is Fixed Asset", "Asset Category", "Brand",
    "Description", "End of Life", "Shelf Life In Days", "Variant Of",
    "Variant Based On", "Safety Stock", "Lead Time in days", "HSN/SAC",
    "Attribute (Variant Attributes)", "Attribute Value (Variant Attributes)"
]

def load_fixed_variants():
    """Load Fixed_Variants.xlsx into a dict: item_code → list of row dicts."""
    wb = openpyxl.load_workbook(FIXED_VARIANTS_PATH, data_only=True)
    ws = wb.active

    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    col = {h: i for i, h in enumerate(header)}

    items = {}       # item_code → [row_dict, ...]  (primary + continuation rows)
    current_code = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        item_code = str(row[col["Item Code"]]).strip() if row[col["Item Code"]] else ""

        if item_code:
            current_code = item_code
            items[current_code] = []

        if current_code is None:
            continue

        row_dict = {}
        for h in HEADERS:
            if h in col:
                val = row[col[h]]
                row_dict[h] = str(val).strip() if val is not None else ""
            else:
                row_dict[h] = ""

        items[current_code].append(row_dict)

    return items

def load_item12_variants():
    """
    Read Item(12).csv and extract groups of rows for each variant item.
    Returns: list of (item_code, [row_dict, ...])
    Groups are (primary_row + blank-code continuation rows).
    """
    groups = []
    current_code = None
    current_rows = []

    with open(ITEM12_PATH, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            item_code = row.get("Item Code", "").strip()
            variant_of = row.get("Variant Of", "").strip()

            if item_code:
                # Save previous group if it was a variant
                if current_code and current_rows:
                    first = current_rows[0]
                    if first.get("Variant Of", "").strip():
                        groups.append((current_code, current_rows))

                current_code = item_code
                current_rows = [row]
            else:
                # Continuation row (blank Item Code)
                current_rows.append(row)

        # Don't forget the last group
        if current_code and current_rows:
            first = current_rows[0]
            if first.get("Variant Of", "").strip():
                groups.append((current_code, current_rows))

    return groups

def build_output_rows(item12_groups, fixed_lookup):
    """
    For each item12 variant group, use Fixed_Variants data if available,
    otherwise fall back to Item(12).csv data.
    Returns list of row dicts ready to write.
    """
    output = []
    stats = {"from_fixed": 0, "from_item12": 0, "skipped": 0}

    for item_code, i12_rows in item12_groups:
        if item_code in fixed_lookup:
            # Use correct data from Fixed_Variants.xlsx
            output.extend(fixed_lookup[item_code])
            stats["from_fixed"] += 1
        else:
            # Item not in Fixed_Variants — use Item(12) data, mapped to HEADERS
            for i, r in enumerate(i12_rows):
                row_dict = {}
                for h in HEADERS:
                    row_dict[h] = r.get(h, "").strip()
                # Blank Item Code for continuation rows
                if i > 0:
                    row_dict["Item Code"] = ""
                output.append(row_dict)
            stats["from_item12"] += 1

    return output, stats

def write_output(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Item12 Variants Corrected"

    # Header row
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)

    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    # Data rows
    alt_fill_a = PatternFill("solid", fgColor="EBF3FB")
    alt_fill_b = PatternFill("solid", fgColor="FFFFFF")
    cont_fill  = PatternFill("solid", fgColor="F5F5F5")

    item_count = 0
    for row_idx, row_dict in enumerate(rows, 2):
        is_primary = bool(row_dict.get("Item Code", "").strip())
        if is_primary:
            item_count += 1
            fill = alt_fill_a if item_count % 2 == 0 else alt_fill_b
        else:
            fill = cont_fill

        for col_idx, h in enumerate(HEADERS, 1):
            val = row_dict.get(h, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val else None)
            cell.fill = fill
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal="left")

    # Column widths
    widths = [30, 18, 18, 8, 8, 8, 14, 14, 30, 18, 12, 20, 14, 10, 12, 12, 30, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    wb.save(OUTPUT_PATH)

def main():
    print("Loading Fixed_Variants.xlsx …")
    fixed_lookup = load_fixed_variants()
    print(f"  Loaded {len(fixed_lookup)} items from Fixed_Variants.xlsx")

    print("Reading Item(12).csv variant groups …")
    item12_groups = load_item12_variants()
    print(f"  Found {len(item12_groups)} variant items in Item(12).csv")

    print("\nVariant items found in Item(12).csv:")
    for code, rows in item12_groups:
        first = rows[0]
        template = first.get("Variant Of", "")
        attrs = [(r.get("Attribute (Variant Attributes)", ""), r.get("Attribute Value (Variant Attributes)", "")) for r in rows]
        in_fixed = "✓ FOUND in Fixed_Variants" if code in fixed_lookup else "✗ NOT in Fixed_Variants (using Item12 data)"
        attr_str = ", ".join(f"{a}={v}" for a,v in attrs if a)
        print(f"  [{in_fixed}]  {code!r:45s}  template={template!r}  attrs=[{attr_str}]")

    print("\nBuilding output rows …")
    output_rows, stats = build_output_rows(item12_groups, fixed_lookup)

    print(f"\nSummary:")
    print(f"  Items from Fixed_Variants (correct attrs): {stats['from_fixed']}")
    print(f"  Items from Item(12).csv (not in Fixed):    {stats['from_item12']}")
    print(f"  Total output rows:                         {len(output_rows)}")

    print(f"\nWriting {OUTPUT_PATH} …")
    write_output(output_rows)
    print("Done! ✓")

if __name__ == "__main__":
    main()
