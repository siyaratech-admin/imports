#!/usr/bin/env python3
"""
Material Library Attribute Fixer
Parses item codes to extract meaningful ERPNext variant attributes.

Run from /home/shubham/kb-bench/apps/:
    python3 "Material Library/fix_attributes.py"

Outputs (in Material Library/):
    Fixed_Item_Attribute_Masters.xlsx  - Item Attribute masters to create first
    Fixed_Parent_Material.xlsx         - Template items with correct attribute names
    Fixed_Variants.xlsx                - Variants multi-row expanded with correct values
"""
import openpyxl, re, os, sys
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ── UOM normalization: map source values → ERPNext standard names ────────────
# Standard ERPNext UOMs that already exist (exact case required):
#   Nos, Kg, Gram, Meter, Litre, Box, Set, Pair, Foot, Inch,
#   Cubic Meter, Square Meter, Square Foot, Cubic Foot
UOM_NORMALIZE = {
    'NOS': 'Nos',
    'KG': 'Kg',
    'BOX': 'Box',
    'SET': 'Set',
    'FEET': 'Foot',          # ERPNext standard is "Foot"
    'CUM': 'Cubic Meter',    # CUM = cubic metre in construction
    # Non-standard UOMs — kept as-is, will be created via prerequisites file
    'BAGS': 'Bag',
    'BRASS': 'Brass',
    'PKT': 'Pkt',
    'RFT': 'RFT',
    'ROLL': 'Roll',
}

# UOMs that need to be created in ERPNext (not in default install)
UOM_TO_CREATE = {
    'Bag':   'Bag',
    'Brass': 'Brass (100 cu.ft)',
    'Pkt':   'Packet',
    'RFT':   'Running Feet',
    'Roll':  'Roll',
}

# Item Groups to create (all under "All Item Groups" root)
ITEM_GROUPS = [
    'Consumable', 'Electricals', 'Fabrication', 'Furniture', 'Hardware',
    'Office Setup', 'Plumbing', 'QA/QC Equipment', 'Raw Material', 'Safety',
    'Spare Parts', 'Stationery', 'Tiles', 'Tools', 'Wooden Shuttering',
]

# ── Brand lists (longer entries first so they match before shorter substrings) ──────
PLUMBING_BRANDS  = ['ASHIRVAD', 'SUPREME', 'DUTRON', 'ASTRAL', 'RAKSHA', 'PRINCE', 'PRINC',
                    'FINOLEX', 'JAIN', 'AKG',
                    'KEROVIT', 'JAQUAR', 'KOHLER', 'JOHNSON', 'SOMANY', 'CERA', 'RAK']
CABLE_BRANDS     = ['POLYCAB', 'LEGRAND', 'PLOYCAB', 'HAVELLS', 'ANCHOR', 'KIRAN', 'L&T', 'RR']
ELEC_BRANDS      = ['LEGRAND', 'L&T', 'POLYCAB', 'PLOYCAB', 'HAVELLS', 'ANCHOR', 'KIRAN', 'DIAMOND', 'ROMA', 'LEGRAND', 'RR']
STEEL_BRANDS     = ['SHREE OM', 'KAMDHENU', 'VIZAG', 'RAJURI', 'GSPL', 'TATA', 'JSW', 'UMA', 'SAIL']
ACC_BRANDS       = ['POWER LITE', 'ECOLITE', 'SIPOREX', 'ULTRATECH', 'GODREJ', 'DLITE']
PAINT_BRANDS     = ['NEROLACK', 'POLYCAB', 'BERGER', 'ASIAN', 'BIRLA', 'DULUX', 'INDIGO', 'JOTUN']
MATERIAL_PREFIXES = ['BRASS', 'GI', 'MS', 'SS', 'RUBBER', 'PVC', 'HDPE']
COLORS = ['STEEL GREY', 'YELLOW', 'WHITE', 'ORANGE', 'BLACK', 'GREEN', 'BLUE', 'RED', 'BROWN', 'GREY', 'IVORY']

# Global registry of attribute→values (built while parsing, used for masters sheet)
attr_registry: dict[str, set] = defaultdict(set)


def reg(attr: str, val: str) -> str:
    """Register a value and return it normalised."""
    v = str(val).strip().upper()
    if v:
        attr_registry[attr].add(v)
    return v


def find_brand(text: str, brand_list: list[str]) -> str | None:
    t = text.upper()
    for b in sorted(brand_list, key=len, reverse=True):
        if b in t:
            return b
    return None


def extract_size_mm(text: str) -> str | None:
    """Return the first size/dimension token ending in MM (e.g. 65MM, 65X50MM, 100X3MM)."""
    m = re.search(r'(\d+(?:\.\d+)?(?:X\d+(?:\.\d+)?){0,2})\s*MM', text.upper())
    return m.group(0).replace(' ', '') if m else None


def get_suffix(code: str, parent: str) -> str:
    """Return the part of the item code that comes after the parent prefix."""
    c, p = code.strip().upper(), parent.strip().upper()
    if c.startswith(p):
        return c[len(p):].strip()
    # Fuzzy: require matched prefix to be at least 50% of parent length (avoids 1-char matches)
    min_match = max(4, len(p) // 2)
    for i in range(len(p) - 1, min_match - 1, -1):
        if c.startswith(p[:i]):
            return c[i:].strip()
    return c   # fallback: return full code unchanged


# ═══════════════════════════════════════════════════════════════════════════════
# Per-group attribute parsers
# Each returns a list of (attribute_name, normalised_value) tuples.
# ═══════════════════════════════════════════════════════════════════════════════

def _last_word_brand(code: str, brand_list: list) -> str | None:
    """Find brand from the LAST word of the item code.
    Needed when brand names also appear in product model names (e.g. 'LK & Jaquar CERA').
    """
    words = code.upper().split()
    for w in reversed(words):
        if find_brand(w, brand_list):
            return find_brand(w, brand_list)
    return None


def parse_plumbing(code: str, parent: str, suffix: str) -> list:
    s = suffix.upper()
    c = code.upper()
    p = parent.strip().upper()

    # For most plumbing items the brand is the last word (most reliable when model names
    # contain other brand names, e.g. "BIB COCK LK & Jaquar CERA").
    brand = _last_word_brand(c, PLUMBING_BRANDS) or find_brand(s, PLUMBING_BRANDS)
    s2 = s.replace(brand, '').strip() if brand else s

    # UPVC END PLUG vs UPVC END CAP (both map to UPVC END PLUG template in source data)
    if p == 'UPVC END PLUG':
        cap_type = 'CAP' if 'END CAP' in c else 'PLUG'
        size = extract_size_mm(s2) or extract_size_mm(s)
        attrs = [('Type', reg('Type', cap_type))]
        if size:
            attrs.append(('Size', reg('Size', size)))
        if brand:
            attrs.append(('Brand', reg('Brand', brand)))
        return attrs

    # TALL PILLER COCK template also contains SELF CLOSING URINAL variants in source data
    if p == 'TALL PILLER COCK':
        item_type = 'URINAL' if 'URINAL' in c else 'PILLAR'
        attrs = [('Type', reg('Type', item_type))]
        if brand:
            attrs.append(('Brand', reg('Brand', brand)))
        return attrs

    # Lab cock with ways: "1WAY JAQUAR", "2WAY CERA"
    m = re.match(r'(\d+)\s*WAY(?:\s+(.+))?', s)
    if m:
        attrs = [('Ways', reg('Ways', m.group(1) + ' WAY'))]
        if m.group(2):
            b = find_brand(m.group(2).strip(), PLUMBING_BRANDS)
            if b:
                attrs.append(('Brand', reg('Brand', b)))
        return attrs

    # BIB COCK / ANGLE COCK: multiple product types under same template → add Type
    if p in ('BIB COCK', 'ANGLE COCK'):
        pidx = c.find(p)
        before = c[:pidx].strip() if pidx >= 0 else ''
        after  = c[pidx + len(p):].strip() if pidx >= 0 else s
        type_kws = [('PRESSMATICS', 'PRESSMATICS'), ('PRESSMATIC', 'PRESSMATIC'),
                    ('TWO WAY', 'TWO WAY'), ('PVC', 'PVC'), ('L&K', 'LK'), ('LK', 'LK')]
        ptype = None
        for k, v in type_kws:
            if k in before or after.upper().startswith(k):
                ptype = v
                break
        attrs = []
        if ptype:
            attrs.append(('Type', reg('Type', ptype)))
        if brand:
            attrs.append(('Brand', reg('Brand', brand)))
        return attrs or [('Brand', reg('Brand', brand))] if brand else [('Size/Type', reg('Size/Type', suffix or code))]

    # PIPELINE O RING: size + manufacturer type (AQUARIUS, SCHWINGSTETTER)
    if 'O RING' in p or 'O RING' in c:
        size = extract_size_mm(s)
        type_m = re.search(r'(?:MM\s*)([A-Z]+)\s*TYPE', s, re.I)
        attrs = []
        if size:
            attrs.append(('Size', reg('Size', size)))
        if type_m:
            attrs.append(('Type', reg('Type', type_m.group(1).strip())))
        elif brand:
            attrs.append(('Brand', reg('Brand', brand)))
        return attrs or [('Size/Type', reg('Size/Type', suffix or code))]

    size = extract_size_mm(s2)
    if not size:
        m = re.search(r'[\d/]+"', s2)
        size = m.group(0) if m else None
    attrs = []
    if size:
        attrs.append(('Size', reg('Size', size)))
    if brand:
        attrs.append(('Brand', reg('Brand', brand)))
    return attrs or [('Size/Type', reg('Size/Type', suffix or code))]


def parse_electricals(code: str, parent: str, suffix: str) -> list:
    s = suffix.upper()

    # MCB: "16 AMP X 2 POLE LEGRAND"
    m = re.match(r'(\d+)\s*AMP\s*X\s*(\d+)\s*POLE\s+(.+)', s)
    if m:
        return [('Amperage', reg('Amperage', m.group(1) + ' AMP')),
                ('Poles',    reg('Poles',    m.group(2) + ' POLE')),
                ('Brand',    reg('Brand',    m.group(3).strip()))]

    # TOP PIN: "6 AMPX2 PIN"
    m = re.match(r'(\d+)\s*AMP\s*X\s*(\d+)\s*PIN', s)
    if m:
        return [('Amperage', reg('Amperage', m.group(1) + ' AMP')),
                ('Pins',     reg('Pins',     m.group(2) + ' PIN'))]

    # Cable: "1 SQMMX2 CORE KIRAN" / "RED 1 SQMM X1 CORE KIRAN" (with optional color prefix)
    # Use re.search so color words before the SQMM section don't block the match
    m = re.search(r'([\d.]+)\s*SQMM\s*X?\s*(\d+)\s*CORE\s+(.+)', s)
    if m:
        color = next((c for c in sorted(COLORS, key=len, reverse=True) if c in s[:m.start()]), None)
        attrs = []
        if color:
            attrs.append(('Color', reg('Color', color)))
        attrs += [('Cross Section', reg('Cross Section', m.group(1) + ' SQMM')),
                  ('Cores',         reg('Cores',         m.group(2) + ' CORE')),
                  ('Brand',         reg('Brand',         m.group(3).strip()))]
        return attrs

    # Industrial socket/plug/top: "16 AMP 3 PIN IP 67" / "32 AMP 5 PIN IP 44"
    m = re.match(r'(\d+)\s*AMP\s+(\d+)\s*PIN\s+(IP\s*\d+)', s)
    if m:
        return [('Amperage',  reg('Amperage',  m.group(1) + ' AMP')),
                ('Pins',      reg('Pins',      m.group(2) + ' PIN')),
                ('IP Rating', reg('IP Rating', m.group(3).replace(' ', '')))]

    # RCCB / CHANGE OVER: "16 AMP LEGRAND" or "100AMP/415V" (amperage + brand/spec)
    m = re.match(r'(\d+)\s*AMP\s+(.+)', s)
    if m:
        return [('Amperage', reg('Amperage', m.group(1) + ' AMP')),
                ('Brand',    reg('Brand',    m.group(2).strip()))]
    # BUS BAR: "100AMP/415V" (no space, dual spec)
    m = re.match(r'(\d+)\s*AMP', s)
    if m:
        return [('Amperage', reg('Amperage', m.group(1) + ' AMP'))]

    # DB BOARD: "IP 67 LEGRAND" / "IP 44 IP 44 LEGRAND"
    m = re.match(r'(IP\s*\d+)', s)
    if m:
        brand = find_brand(s, ELEC_BRANDS)
        attrs = [('IP Rating', reg('IP Rating', m.group(1).replace(' ', '')))]
        if brand:
            attrs.append(('Brand', reg('Brand', brand)))
        return attrs

    # MCB BOX / Junction Box / Metal Box: "4 WAY", "1 WAY DIAMOND", "2 WAY ANCHOR"
    m = re.match(r'(\d+)\s*WAY(?:\s+(.+))?', s)
    if m:
        attrs = [('Ways', reg('Ways', m.group(1) + ' WAY'))]
        if m.group(2):
            b = find_brand(m.group(2).strip(), ELEC_BRANDS)
            if b:
                attrs.append(('Brand', reg('Brand', b)))
        return attrs

    # Bulb: "40 WT"
    m = re.match(r'(\d+)\s*WT', s)
    if m:
        return [('Wattage', reg('Wattage', m.group(1) + ' WT'))]

    # PVC BOARD: "4 MODULAR"
    m = re.match(r'(\d+)\s*MODULAR', s)
    if m:
        return [('Modular', reg('Modular', m.group(1) + ' MODULAR'))]

    # LCB / isolator without AMP: "2 POLE LEGRAND", "3 POLE L&T"
    m = re.match(r'(\d+)\s*POLE\s+(.+)', s)
    if m:
        return [('Poles',  reg('Poles',  m.group(1) + ' POLE')),
                ('Brand',  reg('Brand',  m.group(2).strip()))]

    # Telephone / data cable: "2 PAIR POLYCAB", "5 PAIR POLYCAB"
    m = re.match(r'(\d+)\s*PAIR\s+(.+)', s)
    if m:
        return [('Pairs',  reg('Pairs',  m.group(1) + ' PAIR')),
                ('Brand',  reg('Brand',  m.group(2).strip()))]

    # PVC conduit fittings with brand: "25MM DIAMOND", "25 MM CONDUIT ANCHOR"
    m = re.match(r'([\d.]+)\s*MM(?:\s+CONDUIT)?\s+(.+)', s)
    if m:
        return [('Size',  reg('Size',  m.group(1) + 'MM')),
                ('Brand', reg('Brand', m.group(2).strip()))]

    # Plain size in MM (no brand): "10 MM", "1.5 MM", "20 MM" (flexible pipe, lugs, cassing)
    m = re.match(r'^([\d.]+)\s*MM$', s.strip())
    if m:
        return [('Size', reg('Size', m.group(1) + 'MM'))]

    # Metal switch box / module box with brand: "18 W ROMA", "12 W LEGRAND"
    m = re.match(r'(\d+)\s*W\s+(.+)', s)
    if m:
        return [('Modular', reg('Modular', m.group(1) + ' WAY')),
                ('Brand',   reg('Brand',   m.group(2).strip()))]

    # Plain wattage only: "1500 W", "2000 W" (water heater, geyser)
    m = re.match(r'^(\d+)\s*W$', s.strip())
    if m:
        return [('Wattage', reg('Wattage', m.group(1) + 'W'))]

    # Ceiling fan: "(HIGHSPEED) 24\" CROMPTON" - optional qualifier + inch size + brand
    m = re.search(r'([\d.]+)"\s*(\w[\w&]*)\s*$', s)
    if m:
        return [('Size',  reg('Size',  m.group(1) + '"')),
                ('Brand', reg('Brand', m.group(2).strip()))]

    # Dimension without MM (wall plates, duct size): "8X10", "4X7"
    m = re.match(r'^(\d+X\d+)$', s.strip())
    if m:
        return [('Dimensions', reg('Dimensions', m.group(1)))]

    return [('Size/Type', reg('Size/Type', suffix or code))]


def parse_raw_material(code: str, parent: str, suffix: str) -> list:
    s, p = suffix.upper(), parent.strip().upper()

    if p == 'STEEL':
        m = re.match(r'(\d+)\s*MM\s+(TMT\s+[\d\w-]+)\s+(.+)', s)
        if m:
            brand = find_brand(m.group(3), STEEL_BRANDS) or m.group(3).strip()
            return [('Diameter', reg('Diameter', m.group(1) + 'MM')),
                    ('Grade',    reg('Grade',    m.group(2).strip())),
                    ('Brand',    reg('Brand',    brand))]

    if p == 'RMC':
        m = re.match(r'(M-[\d]+(?:\s*FF)?)', s)
        if m:
            return [('Concrete Grade', reg('Concrete Grade', m.group(1).strip()))]

    if p == 'ACC BLOCK':
        m = re.match(r'(\d+X\d+X\d+)\s*MM\s+(.+)', s)
        if m:
            brand = find_brand(m.group(2), ACC_BRANDS) or m.group(2).strip()
            return [('Dimensions', reg('Dimensions', m.group(1) + 'MM')),
                    ('Brand',      reg('Brand',      brand))]

    if p == 'BRICKS':
        m = re.match(r'(RED|FLY ASH)\s+(\d+")', s)
        if m:
            return [('Brick Type', reg('Brick Type', m.group(1))),
                    ('Brick Size', reg('Brick Size', m.group(2)))]

    if p == 'METAL':
        m = re.match(r'(\d+)MM', s)
        if m:
            return [('Size', reg('Size', m.group(0)))]

    # GRAY KRUB STONE: "500X400X100MM"
    m = re.match(r'(\d+X\d+X\d+MM)', s)
    if m:
        return [('Dimensions', reg('Dimensions', m.group(1)))]

    size = extract_size_mm(s)
    if size:
        return [('Size', reg('Size', size))]

    return [('Size/Type', reg('Size/Type', suffix or code))]


def parse_hardware(code: str, parent: str, suffix: str) -> list:
    c, p, s = code.upper(), parent.strip().upper(), suffix.upper()

    # BINDING WIRE matched from full code (item code has MS/GI prefix before parent)
    m = re.match(r'(MS|GI)\s+BINDING WIRE\s+(\d+)\s*GU?A?GE', c)
    if m:
        return [('Material Type', reg('Material Type', m.group(1))),
                ('Gauge',         reg('Gauge',         m.group(2) + ' GAUGE'))]

    # RCC PIPE: "RCC PIPE 900MM/NP2/NP3"
    m = re.match(r'RCC PIPE\s+(\d+)MM(?:/(.+))?', c)
    if m:
        attrs = [('Diameter', reg('Diameter', m.group(1) + 'MM'))]
        if m.group(2):
            attrs.append(('Grade', reg('Grade', m.group(2).rstrip('/'))))
        return attrs

    # COVER BLOCKS: PLASTIC, SQUARE 20MM, ROUND 50MM ZF (trailing qualifier kept in size)
    if p == 'COVER BLOCKS':
        shape_m = re.search(r'\b(PLASTIC|SQUARE|ROUND)\b', c)
        # Capture size + optional trailing qualifier (ZF, ZR, ZS) to keep combos unique
        size_m  = re.search(r'(\d+(?:/\d+)?)\s*MM(?:\s+([A-Z]{1,4}))?', c)
        attrs = []
        if shape_m:
            attrs.append(('Shape', reg('Shape', shape_m.group(1))))
        if size_m:
            size_val = size_m.group(1).replace('/', 'X') + 'MM'
            if size_m.group(2):
                size_val += size_m.group(2)
            attrs.append(('Size', reg('Size', size_val)))
        if attrs:
            return attrs

    # Water paper / sandpaper: "3MM 120 NO" → size + grit number
    m = re.match(r'(\d+(?:\.\d+)?)\s*MM\s+(\d+)\s*NO\b', s, re.I)
    if m:
        return [('Size', reg('Size', m.group(1) + 'MM')),
                ('Grit', reg('Grit', m.group(2) + ' NUMBER'))]

    # Cutting wheel: size + brand (BOSCH/HITACHI) or type prefix (TILES/PLY)
    if 'CUTTING WHEEL' in c:
        size = extract_size_mm(c)
        type_m = re.match(r'^(TILES|PLY|WOOD|STONE|METAL)\b', c)
        brand_m = find_brand(c, ['BOSCH', 'HITACHI', 'MAKITA', 'DEWALT'])
        attrs = []
        if size:
            attrs.append(('Size', reg('Size', size)))
        if type_m:
            attrs.append(('Type', reg('Type', type_m.group(1))))
        elif brand_m:
            attrs.append(('Brand', reg('Brand', brand_m)))
        if attrs:
            return attrs

    # WOODEN HANDLE: "FOR HAMMER" / "FOR SPADE"
    if p == 'WOODEN HANDLE':
        m = re.match(r'FOR\s+(.+)', s)
        if m:
            return [('Fits', reg('Fits', m.group(1).strip()))]

    # MS HANDLE vs MS HINGES: source data lumps both under MS HANDLE template
    if p == 'MS HANDLE':
        item_type = 'HINGES' if ('HINGES' in c or 'HINGE' in c) else 'HANDLE'
        size = extract_size_mm(s)
        if not size:
            m2 = re.search(r'([\d.]+")', s)
            size = m2.group(1) if m2 else None
        attrs = [('Material Type', reg('Material Type', 'MS')),
                 ('Type',          reg('Type',          item_type))]
        if size:
            attrs.append(('Size', reg('Size', size)))
        return attrs

    # GI/MS material prefix before parent name (e.g. GI RIDGE CAP, MS PATRA)
    m = re.match(r'^(GI|MS)\b', c)
    if m:
        mat = m.group(1)
        rest = c[len(mat):].strip()
        if p.upper() in rest:
            rest = rest.replace(p.upper(), '').strip()
        attrs = [('Material Type', reg('Material Type', mat))]
        size = extract_size_mm(rest)
        if size:
            attrs.append(('Size', reg('Size', size)))
        else:
            # inch dimensions: 2"X2", 4"X4"
            m2 = re.search(r'([\d."]+"X[\d.]+")', rest)
            if m2:
                attrs.append(('Dimensions', reg('Dimensions', m2.group(1))))
            # single inch: 8", 10"
            elif re.search(r'[\d.]+"', rest):
                m2 = re.search(r'([\d.]+")', rest)
                attrs.append(('Size', reg('Size', m2.group(1))))
            # foot dimensions: 8'X4'
            elif re.search(r"\d+'X\d+'", rest):
                m2 = re.search(r"(\d+'X\d+')", rest)
                attrs.append(('Dimensions', reg('Dimensions', m2.group(1))))
            # single foot: 8', 10', 12'
            elif re.search(r"\d+'", rest):
                m2 = re.search(r"(\d+)'", rest)
                attrs.append(('Length', reg('Length', m2.group(1) + "FT")))
            # XxY without unit: 300X300, 600X600
            else:
                m2 = re.search(r'(\d+X\d+)', rest)
                if m2:
                    attrs.append(('Dimensions', reg('Dimensions', m2.group(1))))
        return attrs

    # Generic size in MM
    size = extract_size_mm(s)
    if size:
        return [('Size', reg('Size', size))]

    # Load capacity in TON: chain block, hoist
    m = re.match(r'(\d+(?:\.\d+)?)\s*TON', s)
    if m:
        return [('Capacity', reg('Capacity', m.group(1) + ' TON'))]

    # Nail / fastener size: "2.5\"X10", "3\"X12" (length×gauge)
    m = re.search(r'[\d.]+"\s*X\s*\d+', s)
    if m:
        return [('Size', reg('Size', m.group(0).replace(' ', '')))]

    # Tarpaulin / sheet feet dimensions: "9'X12'"
    m = re.search(r"[\d.]+'\s*X\s*[\d.]+'", s)
    if m:
        return [('Dimensions', reg('Dimensions', m.group(0).replace(' ', '')))]

    # Capacity in litres: "500 LTR", "10000 LTR"
    m = re.match(r'(\d+)\s*LTR', s)
    if m:
        return [('Capacity', reg('Capacity', m.group(1) + ' LTR'))]

    # Screw/bolt/grip dimension without MM: "75X8", "25X6" (length×gauge)
    m = re.match(r'^(\d+X\d+(?:/\d+)?)$', s.strip())
    if m:
        return [('Size', reg('Size', m.group(1)))]

    # Single inch size: "8\"", "10\"", "2.5\"" (door hardware, hinges, handles)
    m = re.match(r'^([\d.]+)"$', s.strip())
    if m:
        return [('Size', reg('Size', m.group(0)))]

    # Single foot length: "1.5'", "2'", "4'" (wooden tools, bali, planks)
    m = re.match(r"^([\d.]+)'$", s.strip())
    if m:
        return [('Length', reg('Length', m.group(0)))]

    # Sandpaper / abrasive grit: "0 NUMBER", "80 NUMBER", "120 NUMBER"
    m = re.match(r'(\d+)\s*NUMBER', s)
    if m:
        return [('Grit', reg('Grit', m.group(1) + ' NUMBER'))]

    # Broom / brush type: if suffix is the item type keyword (HARD, SOFT, etc.)
    if p == 'BROOM' or 'BROOM' in p:
        return [('Type', reg('Type', (suffix or s).strip()))]

    # Material prefix before parent for mesh/jali/net items: "FIBER JALI", "PVC MESH JALI"
    mat_m2 = re.match(r'^(FIBER|HDPE|PP|NYLON|PVC)\b', c)
    if mat_m2:
        return [('Material Type', reg('Material Type', mat_m2.group(1)))]

    # Material type keyword (only if suffix itself IS the material)
    if s.strip() in ('METAL', 'PVC', 'GI', 'MS', 'RUBBER', 'COTTON', 'LEATHER', 'WOODEN', 'SS', 'FIBER'):
        return [('Material Type', reg('Material Type', s.strip()))]

    return [('Size/Type', reg('Size/Type', suffix or code))]


def parse_consumable(code: str, parent: str, suffix: str) -> list:
    c, p, s = code.upper(), parent.strip().upper(), suffix.upper()

    if p == 'FOAMSHEET':
        m = re.match(r'(\d+)\s*MM', s)
        if m:
            return [('Thickness', reg('Thickness', m.group(1) + 'MM'))]

    if p == 'COTTON ROPE':
        m = re.match(r'(\d+)\s*MM', s)
        if m:
            return [('Diameter', reg('Diameter', m.group(1) + 'MM'))]

    if p == 'MAT':
        if 'RUBBER' in c:
            return [('Material Type', reg('Material Type', 'RUBBER'))]
        m = re.search(r"[\d.]+\'\s*X\s*[\d.]+'", s)
        if m:
            return [('Dimensions', reg('Dimensions', m.group(0).replace(' ', '')))]

    if p == 'MATRIX BED':
        m = re.search(r"[\d.]+\'\s*X\s*[\d.]+'", s)
        if m:
            return [('Dimensions', reg('Dimensions', m.group(0).replace(' ', '')))]

    if p == 'PLASTIC SHEET':
        if 'BLACK' in s:
            return [('Color', reg('Color', 'BLACK'))]
        return [('Type', reg('Type', s.strip()))]

    # Paint / coating items: OIL PAINT, SPRAY PAINT, PRIMER, APEX, WEATHERSHIELD
    paint_parents = ['OIL PAINT', 'SPRAY PAINT', 'PRIMER', 'APEX', 'WEATHERSHIELD',
                     'ENAMEL PAINT', 'EMULSION PAINT', 'BITUMEN PAINT', 'EPOXY PAINT']
    if any(pp in c for pp in paint_parents) or any(pp in p for pp in ['PAINT', 'PRIMER', 'APEX']):
        color = next((col for col in sorted(COLORS, key=len, reverse=True) if col in s), None)
        # APEX PAINT: brand is usually the LAST word (e.g. "K107 BIRLA" → BIRLA, not the product ASIAN)
        if 'APEX' in p or ('APEX' in c and p not in ('OIL PAINT',)):
            words = c.split()
            brand = next((w for w in reversed(words) if w in PAINT_BRANDS), None) or find_brand(s, PAINT_BRANDS)
        else:
            brand = find_brand(s, PAINT_BRANDS)
        attrs = []
        # OIL BISON ACRYLIC DISTEMPER items are incorrectly under OIL PAINT template in source data;
        # add Paint Type to make their attribute combo distinct from OIL PAINT items.
        if p == 'OIL PAINT' and 'BISON' in c:
            attrs.append(('Paint Type', reg('Paint Type', 'BISON')))
        if color:
            attrs.append(('Color', reg('Color', color)))
        if brand:
            attrs.append(('Brand', reg('Brand', brand)))
        if attrs:
            return attrs

    # Tile grout / tile adhesive: Color variants (may also have grade like SF)
    if any(t in p for t in ['GROUT', 'ADHESIVE', 'PUTTY']):
        color = next((col for col in sorted(COLORS, key=len, reverse=True) if col in s), None)
        brand = find_brand(s, PAINT_BRANDS)
        # Grade: short token at start before brand, e.g. "SF" in "SF BERGER"
        grade_m = re.match(r'^([A-Z]{1,4}\+?)\b', s)
        grade = grade_m.group(1) if grade_m and grade_m.group(1) not in ['FOR'] else None
        attrs = []
        if color:
            attrs.append(('Color', reg('Color', color)))
        if grade and not color:
            attrs.append(('Grade', reg('Grade', grade)))
        if brand:
            attrs.append(('Brand', reg('Brand', brand)))
        if attrs:
            return attrs

    # Cement types (WHITE CEMENT, OPC, PPC)
    if p == 'CEMENT':
        return [('Cement Type', reg('Cement Type', (suffix or code).strip().upper()))]

    # Lubricants: Engine Oil, Gear Oil, Hydraulic Oil → Viscosity Grade
    if any(x in p for x in ['ENGINE OIL', 'GEAR OIL', 'HYDRAULIC OIL', 'COMPRESSOR OIL']):
        grade = (suffix or s).strip()
        return [('Viscosity Grade', reg('Viscosity Grade', grade))]

    # Grease: Grade (AP3, EP00, etc.)
    if 'GREASE' in p:
        return [('Grade', reg('Grade', (suffix or s).strip()))]

    # Thinner: Type (G P = General Purpose, N C = Nitrocellulose)
    if 'THINNER' in p:
        return [('Thinner Type', reg('Thinner Type', (suffix or s).strip()))]

    # Chemical admixtures / waterproofing (DR FIXIT, NITO BOND, SIKA, FOSROC)
    if any(x in c for x in ['DR FIXIT', 'DR.FIXIT', 'NITO BOND', 'SIKA', 'FOSROC', 'PIDILITE', 'MYK', 'FIXIT']):
        return [('Product Type', reg('Product Type', (suffix or s).strip()))]

    # 3D dimensions like 30X30X200MM — must check before the 2D fallback below
    size3d = extract_size_mm(s)
    if size3d and 'X' in size3d:
        return [('Dimensions', reg('Dimensions', size3d))]

    # Lintel / structural items with inch-feet dimensions: "8\"X3'"
    m = re.search(r'[\d.]+"\s*X\s*[\d.]+"?\'?|[\d.]+\'\s*X\s*[\d.]+"', s)
    if not m:
        m = re.search(r'[\d.]+["\']?\s*X\s*[\d.]+["\']?', s)
    if m:
        return [('Dimensions', reg('Dimensions', m.group(0).replace(' ', '')))]

    # Cover blocks (group may be CONSUMABLE in variants file)
    if 'COVER BLOCK' in c:
        shape_m = re.search(r'\b(PLASTIC|SQUARE|ROUND)\b', c)
        size_m  = re.search(r'(\d+(?:/\d+)?)\s*MM(?:\s+([A-Z]{1,4}))?', c)
        attrs = []
        if shape_m:
            attrs.append(('Shape', reg('Shape', shape_m.group(1))))
        if size_m:
            size_val = size_m.group(1).replace('/', 'X') + 'MM'
            if size_m.group(2):
                size_val += size_m.group(2)
            attrs.append(('Size', reg('Size', size_val)))
        if attrs:
            return attrs

    # Capacity in litres: "500 LTR", "10000 LTR" (water tanks etc.)
    m = re.match(r'(\d+)\s*LTR', s)
    if m:
        return [('Capacity', reg('Capacity', m.group(1) + ' LTR'))]

    # Acid / chemical type (WHITE CAT ACID, etc.)
    if any(x in p for x in ['ACID', 'SOLVENT', 'FLUX']):
        return [('Chemical Type', reg('Chemical Type', (suffix or s).strip()))]

    size = extract_size_mm(s)
    if size:
        return [('Size', reg('Size', size))]

    # Inch/feet dimensions for structural consumables (LINTEL 8"X3')
    m = re.search(r'[\d.]+["\']?\s*X\s*[\d.]+["\']?', s)
    if m:
        return [('Dimensions', reg('Dimensions', m.group(0).replace(' ', '')))]

    return [('Size/Type', reg('Size/Type', suffix or code))]


def parse_fabrication(code: str, parent: str, suffix: str) -> list:
    s, c = suffix.upper(), code.upper()
    # 3D: 100X100X4MM or 10MMX10MMX2MM
    m = re.search(r'(\d+(?:MM)?X\d+(?:MM)?X\d+(?:MM)?)', s)
    if m:
        return [('Dimensions', reg('Dimensions', m.group(1).replace('MM', '') + 'MM'))]
    # 2D: 75X6MM
    m = re.match(r'(\d+(?:\.\d+)?X\d+(?:\.\d+)?)\s*MM', s)
    if m:
        return [('Dimensions', reg('Dimensions', m.group(1) + 'MM'))]
    # Product type + size when suffix has 'BAR' / 'PIPE' / 'CHANNEL': "BAR 5MM"
    m = re.match(r'(BAR|PIPE|CHANNEL|BEAM|JOIST)\s*(\d+\s*MM)', s)
    if m:
        return [('Type',  reg('Type',  m.group(1))),
                ('Size',  reg('Size',  m.group(2).replace(' ', '')))]
    # MS prefix items: "MS ROUND BAR 08 MM" - extract size from anywhere
    size = extract_size_mm(s) or extract_size_mm(c)
    if size:
        return [('Size', reg('Size', size))]
    return [('Size/Type', reg('Size/Type', suffix or code))]


def parse_safety(code: str, parent: str, suffix: str) -> list:
    c, p, s = code.upper(), parent.strip().upper(), suffix.upper()

    # Helmet/Jacket color
    colors = ['STEEL GREY', 'YELLOW', 'WHITE', 'ORANGE', 'BLACK', 'GREEN', 'BLUE', 'RED']
    for col in sorted(colors, key=len, reverse=True):
        if col in s:
            attrs = [('Color', reg('Color', col))]
            if 'M/C' in s:
                attrs.append(('Grade', reg('Grade', 'M/C')))
            elif 'F/C' in s:
                attrs.append(('Grade', reg('Grade', 'F/C')))
            return attrs

    # Hand Gloves material (handles typo RUBBUR)
    mat_m = re.match(r'^(RUBB(?:UR|ER)|COTTON|LEATHER)', s)
    if mat_m:
        mat = 'RUBBER' if 'RUBB' in mat_m.group(1) else mat_m.group(1)
        return [('Material Type', reg('Material Type', mat))]

    # Glove material from full item code prefix
    mat2 = re.match(r'^(RUBB(?:UR|ER)|COTTON|LEATHER|NYLON)', c)
    if mat2:
        mat = 'RUBBER' if 'RUBB' in mat2.group(1) else mat2.group(1)
        return [('Material Type', reg('Material Type', mat))]

    # Safety shoes type
    if p == 'SAFETY SHOES':
        return [('Type', reg('Type', (suffix or s).strip()))]

    # Rope / cable size in MM
    size = extract_size_mm(s)
    if size:
        return [('Size', reg('Size', size))]

    # All other safety items: use as Type
    val = re.sub(r'^[\-\s]+', '', (suffix or code)).strip().upper()
    if val:
        return [('Type', reg('Type', val))]

    return [('Size/Type', reg('Size/Type', (suffix or code).strip().upper()))]


def parse_tiles(code: str, parent: str, suffix: str) -> list:
    s, p = suffix.upper(), parent.strip().upper()
    colors = ['STEEL GREY', 'BLACK', 'WHITE', 'BROWN', 'IVORY', 'YELLOW', 'GREEN', 'BLUE', 'RED']

    # Granite/Marble/Kadappa: detect sub-type (T, Z) and finish (LEATHER, COFF) BEFORE color check
    if p in ('GRANITE', 'KADAPPA', 'MARBLE'):
        type_m = re.match(r'^([A-Z])\s+', s)   # single-letter prefix: T, Z
        ptype  = type_m.group(1) if type_m else None
        finish = 'LEATHER' if re.search(r'LETH?ER', s, re.I) else None
        color  = next((col for col in sorted(colors, key=len, reverse=True) if col in s), None)
        attrs  = []
        if ptype:
            attrs.append(('Type', reg('Type', ptype)))
        elif finish:
            attrs.append(('Finish', reg('Finish', finish)))
        elif color:
            # Detect shade/prefix text before the color (e.g. "COFF" in "COFF BROWN")
            before_color = s[:s.find(color)].strip()
            if before_color:
                attrs.append(('Finish', reg('Finish', before_color)))
        if color:
            attrs.append(('Color', reg('Color', color)))
        return attrs or [('Finish/Color', reg('Finish/Color', s.strip()))]

    for col in sorted(colors, key=len, reverse=True):
        if col in s:
            # For TILES: also detect model/brand prefix to distinguish items with same color
            before_col = s[:s.find(col)].strip()
            prefix_m = re.match(r'^([A-Z]+)\b', before_col) if before_col else None
            attrs = [('Color', reg('Color', col))]
            if prefix_m and prefix_m.group(1) not in ('DARK', 'LIGHT', 'DEEP', 'ANTI'):
                attrs.insert(0, ('Series', reg('Series', prefix_m.group(1))))
            return attrs
    m = re.search(r'(\d+X\d+)\s*MM', s)
    if m:
        return [('Dimensions', reg('Dimensions', m.group(1) + 'MM'))]
    return [('Size/Type', reg('Size/Type', suffix or code))]


def parse_tools(code: str, parent: str, suffix: str) -> list:
    c, s = code.upper(), suffix.upper()

    # GI/MS prefix before parent (e.g. GI CONCRETE TRAY 600X600MM, MS SABBAL 5')
    m = re.match(r'^(GI|MS)\b', c)
    if m:
        mat = m.group(1)
        rest = c[len(mat):].strip()
        if parent.upper() in rest:
            rest = rest.replace(parent.upper(), '').strip()
        attrs = [('Material Type', reg('Material Type', mat))]
        size = extract_size_mm(rest)
        if size:
            attrs.append(('Dimensions', reg('Dimensions', size)) if 'X' in size else ('Size', reg('Size', size)))
        else:
            m2 = re.search(r"[\d.'\"]+X[\d.'\"]+", rest)
            if m2:
                attrs.append(('Dimensions', reg('Dimensions', m2.group(0))))
            elif re.search(r"\d+'", rest):
                m2 = re.search(r"(\d+)'", rest)
                attrs.append(('Length', reg('Length', m2.group(1) + "FT")))
            elif re.search(r'[\d.]+"', rest):
                m2 = re.search(r'([\d.]+")', rest)
                attrs.append(('Size', reg('Size', m2.group(1))))
        return attrs
    # Portable toilet / site cabin in TOOLS group: RENTED vs OWNED + dimensions
    if 'PORTABLE TOILET' in code.upper() or 'TOILET BLOCK' in code.upper():
        ownership = 'RENTED' if 'RENT' in code.upper() else 'OWNED'
        dim_m = re.search(r"[\d.']+X[\d.']+", s)
        attrs = [('Ownership', reg('Ownership', ownership))]
        if dim_m:
            attrs.append(('Dimensions', reg('Dimensions', dim_m.group(0).replace(' ', ''))))
        return attrs

    size = extract_size_mm(s)
    if size:
        return [('Size', reg('Size', size))]
    m = re.search(r"[\d.']+X[\d.']+", s)
    if m:
        return [('Dimensions', reg('Dimensions', m.group(0)))]
    return [('Size/Type', reg('Size/Type', suffix or code))]


def parse_generic(code: str, parent: str, suffix: str) -> list:
    c = code.upper()
    s = suffix.upper() if suffix else c
    p = parent.strip().upper()

    # Battery: "50 AH"
    m = re.match(r'(\d+)\s*AH', s)
    if m:
        return [('Capacity', reg('Capacity', m.group(1) + ' AH'))]

    # Carbon brush machine type
    if p == 'CARBON BRUSH':
        return [('Machine Type', reg('Machine Type', (suffix or code).strip().upper()))]

    # Tyre / Tyre Kit: position (FRONT/REAR) + dimensions + drive type (2WD/4WD/HD)
    if 'TYRE' in p or 'TYRE' in code.upper():
        c_up = code.upper()
        position = 'REAR' if 'REAR' in c_up else 'FRONT' if 'FRONT' in c_up else None
        drive_m  = re.search(r'\b(\d+WD|HD)\b', c_up, re.I)
        dim_m    = re.search(r'(\d+X\d+)', s)
        attrs = []
        if position:
            attrs.append(('Position', reg('Position', position)))
        if dim_m:
            attrs.append(('Dimensions', reg('Dimensions', dim_m.group(1))))
        if drive_m:
            attrs.append(('Drive Type', reg('Drive Type', drive_m.group(1).upper())))
        if attrs:
            return attrs

    # Pipeline O-ring / spare part O-ring: size + manufacturer type
    if 'O RING' in code.upper():
        size    = extract_size_mm(s)
        type_m  = re.search(r'(?:MM\s*)([A-Z]{3,})\s*TYPE', s, re.I)
        attrs   = []
        if size:
            attrs.append(('Size', reg('Size', size)))
        if type_m:
            attrs.append(('Type', reg('Type', type_m.group(1).strip())))
        if attrs:
            return attrs

    # Portable toilet / site cabin: RENTED vs owned
    if 'PORTABLE TOILET' in code.upper() or 'TOILET BLOCK' in code.upper():
        ownership = 'RENTED' if 'RENT' in code.upper() else 'OWNED'
        dim_m = re.search(r"[\d.']+X[\d.']+", s)
        attrs = [('Ownership', reg('Ownership', ownership))]
        if dim_m:
            attrs.append(('Dimensions', reg('Dimensions', dim_m.group(0).replace(' ', ''))))
        return attrs

    # Stationery pages: "200 PAGE"
    m = re.match(r'(\d+)\s*PAGE', s)
    if m:
        return [('Pages', reg('Pages', m.group(1) + ' PAGE'))]

    # Paper size: A3, A4
    m = re.match(r'^(A\d)\b', s)
    if m:
        return [('Paper Size', reg('Paper Size', m.group(1)))]

    # Wire rope length: "6 MTR"
    m = re.match(r'(\d+)\s*MTR', s)
    if m:
        return [('Length', reg('Length', m.group(1) + ' MTR'))]

    # Sheet with thickness: "8'X4'X18MM" — thickness is the differentiator across plywood sizes
    m = re.search(r"'\s*X\s*[\d.]+'\s*X\s*(\d+(?:\.\d+)?)\s*MM", s)
    if m:
        return [('Thickness', reg('Thickness', m.group(1) + "MM"))]

    # Wooden bali length: "16'"
    m = re.match(r"(\d+)'", s)
    if m:
        return [('Length', reg('Length', m.group(1) + "FT"))]

    # USB / Wireless
    for t in ['USB', 'WIRELESS']:
        if t in s:
            return [('Connection Type', reg('Connection Type', t))]

    # Material prefix before parent name: "GI SIEVES SET", "BRASS SIEVES SET 20/75MM"
    mat_m = re.match(r'^(BRASS|GI|MS|SS|RUBBER|HDPE|PVC)\b', c)
    if mat_m:
        mat = mat_m.group(1)
        rest = c[len(mat):].strip()
        if p in rest:
            rest = rest.replace(p, '').strip()
        size = extract_size_mm(rest)
        m_size = re.search(r'(\d+/\d+)MM', rest) if not size else None
        attrs = [('Material Type', reg('Material Type', mat))]
        if size:
            attrs.append(('Size', reg('Size', size)))
        elif m_size:
            attrs.append(('Size', reg('Size', m_size.group(0))))
        return attrs

    # QA/QC cube mould: 150X150X150MM
    dim3 = re.match(r'(\d+X\d+X\d+\s*MM)', s, re.I)
    if dim3:
        return [('Dimensions', reg('Dimensions', dim3.group(1).replace(' ', '')))]

    # Furniture / door size: "6.75'X2.25'" / "88"X30""
    m = re.search(r'[\d.]+["\']?\s*X\s*[\d.]+["\']?', s)
    if m:
        return [('Dimensions', reg('Dimensions', m.group(0).replace(' ', '')))]

    size = extract_size_mm(s)
    if size:
        return [('Size', reg('Size', size))]

    # Amperage without brand: "32 AMP", "63AMP" (fuses, breakers in spare parts)
    m = re.match(r'(\d+)\s*AMP', s)
    if m:
        return [('Amperage', reg('Amperage', m.group(1) + ' AMP'))]

    # Portable toilet / site cabin: RENTED vs owned
    if 'PORTABLE TOILET' in code.upper() or 'TOILET BLOCK' in code.upper():
        ownership = 'RENTED' if 'RENT' in code.upper() else 'OWNED'
        dim_m = re.search(r"[\d.']+X[\d.']+", s)
        attrs = [('Ownership', reg('Ownership', ownership))]
        if dim_m:
            attrs.append(('Dimensions', reg('Dimensions', dim_m.group(0).replace(' ', ''))))
        return attrs

    # Bearing / part by number: "608 NO", "6201 NO"
    m = re.match(r'(\d+)\s*NO\b', s)
    if m:
        return [('Part Number', reg('Part Number', m.group(1)))]

    # V-belt / drive-belt size: "A-69", "B-39"
    m = re.match(r'^([A-Z]-\d+)$', s.strip())
    if m:
        return [('Part Number', reg('Part Number', m.group(1)))]

    # Armature / spare part fitted to machine: "FOR GRINDER 4\"", "FOR BOSCH BREAKER 5 KG"
    m = re.match(r'^FOR\s+(.+)', s)
    if m:
        return [('Machine Type', reg('Machine Type', m.group(1).strip()))]

    # Non-numeric suffix = product type name (furniture, equipment sub-type)
    clean = re.sub(r'^[\-\s]+', '', s).strip()
    if clean and not re.search(r'\d', clean):
        return [('Type', reg('Type', clean))]

    return [('Size/Type', reg('Size/Type', (suffix or code).strip().upper()))]


def parse_attributes(item_code: str, variant_of: str, item_group: str) -> list:
    """Return list of (attribute_name, value) for a variant item."""
    if not variant_of:
        return []
    parent = variant_of.strip()
    code   = (item_code or '').strip()
    suffix = get_suffix(code, parent)
    ig     = (item_group or '').strip().upper()

    dispatch = {
        'PLUMBING':    parse_plumbing,
        'ELECTRICALS': parse_electricals,
        'RAW MATERIAL':parse_raw_material,
        'HARDWARE':    parse_hardware,
        'CONSUMABLE':  parse_consumable,
        'FABRICATION': parse_fabrication,
        'SAFETY':      parse_safety,
        'TILES':       parse_tiles,
        'TOOLS':       parse_tools,
    }
    result = dispatch.get(ig, parse_generic)(code, parent, suffix)
    return [(a, v) for a, v in result if v]


# ═══════════════════════════════════════════════════════════════════════════════
# Spreadsheet helpers
# ═══════════════════════════════════════════════════════════════════════════════

HDR_FILL  = PatternFill('solid', fgColor='1F4E79')
HDR_FONT  = Font(bold=True, color='FFFFFF', size=10)
ALT_FILL  = PatternFill('solid', fgColor='D9E1F2')
WARN_FILL = PatternFill('solid', fgColor='FFF2CC')


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill  = HDR_FILL
        cell.font  = HDR_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)


def freeze_and_filter(ws, row=1):
    ws.freeze_panes = ws.cell(row + 1, 1)
    ws.auto_filter.ref = ws.dimensions


def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ═══════════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parent_path  = os.path.join(BASE_DIR, 'Demo_Parent_Material.xlsx')
    variant_path = os.path.join(BASE_DIR, 'Material_Variants.xlsx')

    print('Reading source files…')
    wb_parent  = openpyxl.load_workbook(parent_path,  read_only=True)
    wb_variant = openpyxl.load_workbook(variant_path, read_only=True)

    parent_rows  = list(wb_parent['Sheet1'].iter_rows(values_only=True))
    variant_rows = list(wb_variant['Sheet1'].iter_rows(values_only=True))

    p_headers = parent_rows[0]
    v_headers = variant_rows[0]

    # Build parent → item_group mapping
    parent_map: dict[str, dict] = {}
    for row in parent_rows[1:]:
        d = dict(zip(p_headers, row))
        if d.get('Item Code'):
            parent_map[d['Item Code'].strip()] = d

    print(f'  Parents: {len(parent_map)}')
    print(f'  Variants: {len(variant_rows) - 1}')

    # ── Parse all variants ────────────────────────────────────────────────────
    # variant_parsed: list of (orig_row_dict, attrs_list, parse_status)
    variant_parsed = []
    fallback_count = 0

    for row in variant_rows[1:]:
        d = dict(zip(v_headers, row))
        item_code  = (d.get('Item Code') or '').strip()
        variant_of = (d.get('Variant Of') or '').strip()

        if not variant_of:
            variant_parsed.append((d, [], 'NO_VARIANT_OF'))
            continue

        # Use variant's Item Group first; fall back to parent's group
        item_group = (d.get('Item Group') or
                      (parent_map.get(variant_of, {}).get('Item Group') or '')).strip()

        attrs = parse_attributes(item_code, variant_of, item_group)

        # If the variant's group gave only a fallback, retry with the parent's group
        is_fallback = any(a == 'Size/Type' for a, _ in attrs)
        if is_fallback:
            parent_group = (parent_map.get(variant_of, {}).get('Item Group') or '').strip()
            if parent_group and parent_group.upper() != item_group.upper():
                attrs2 = parse_attributes(item_code, variant_of, parent_group)
                if attrs2 and not any(a == 'Size/Type' for a, _ in attrs2):
                    attrs = attrs2

        if not attrs:
            attrs = [('Size/Type', (item_code or '').strip())]

        is_fallback = any(a == 'Size/Type' for a, _ in attrs)
        if is_fallback:
            fallback_count += 1
        status = 'REVIEW' if is_fallback else 'OK'
        variant_parsed.append((d, attrs, status))

    # ── Determine attributes per parent (from its variants) ──────────────────
    parent_attrs: dict[str, list[str]] = defaultdict(list)
    for d, attrs, status in variant_parsed:
        variant_of = (d.get('Variant Of') or '').strip()
        if not variant_of:
            continue
        for attr_name, _ in attrs:
            if attr_name not in parent_attrs[variant_of]:
                parent_attrs[variant_of].append(attr_name)

    # ── Write File 1: Item Attribute Masters ──────────────────────────────────
    print('\nWriting Fixed_Item_Attribute_Masters.xlsx…')
    wb1 = openpyxl.Workbook()

    NUMERIC_ATTRS = {'Size', 'Diameter', 'Thickness', 'Gauge', 'Amperage', 'Poles',
                     'Ways', 'Pins', 'Wattage', 'Modular', 'Capacity', 'Cores'}

    sorted_attrs = sorted(attr_registry.keys())

    # ── Sheet 1: ERPNext import format (parent + child rows combined) ──────────
    # Column headers must match ERPNext field labels exactly:
    #   "Attribute Name"                          → Item Attribute.attribute_name
    #   "Numeric Values"                          → Item Attribute.numeric_values
    #   "Attribute Value (Item Attribute Values)" → Item Attribute Value.attribute_value
    ws_import = wb1.active
    ws_import.title = 'ERPNext Import'
    ws_import.append([
        'Attribute Name',
        'Numeric Values',
        'Attribute Value (Item Attribute Values)',
        'Abbreviation (Item Attribute Values)',
    ])
    style_header(ws_import)

    def make_abbr(val: str, used: set) -> str:
        """Generate a unique abbreviation ≤10 chars for an attribute value."""
        # Strip spaces and special chars, uppercase, max 10 chars
        base = re.sub(r'[^A-Z0-9]', '', val.upper())[:10]
        if not base:
            base = 'VAL'
        candidate = base
        counter = 2
        while candidate in used:
            suffix = str(counter)
            candidate = base[:10 - len(suffix)] + suffix
            counter += 1
        used.add(candidate)
        return candidate

    row_idx = 2
    for attr in sorted_attrs:
        vals = sorted(attr_registry[attr])
        used_abbrs: set = set()
        first = True
        for val in vals:
            abbr = make_abbr(val, used_abbrs)
            if first:
                ws_import.append([attr, 0, val, abbr])
                first = False
            else:
                # Continuation rows: blank parent columns, only child value + abbr
                ws_import.append(['', '', val, abbr])
            if row_idx % 2 == 0:
                for cell in ws_import[row_idx]:
                    cell.fill = ALT_FILL
            row_idx += 1

    set_col_widths(ws_import, {'A': 26, 'B': 16, 'C': 36, 'D': 20})
    freeze_and_filter(ws_import)

    # ── Sheet 2: Summary (human-readable reference, NOT for import) ────────────
    ws_sum = wb1.create_sheet('Summary (reference only)')
    ws_sum.append(['Attribute Name', 'Numeric Values', 'Value Count', 'All Values'])
    style_header(ws_sum)
    for i, attr in enumerate(sorted_attrs, 2):
        vals = sorted(attr_registry[attr])
        ws_sum.append([attr, 0, len(vals), ', '.join(vals)])
        if i % 2 == 0:
            for cell in ws_sum[i]:
                cell.fill = ALT_FILL
    set_col_widths(ws_sum, {'A': 26, 'B': 16, 'C': 14, 'D': 120})

    wb1.save(os.path.join(BASE_DIR, 'Fixed_Item_Attribute_Masters.xlsx'))

    # ── Write File 1b: Prerequisites (Item Groups, UOMs, HSN Codes) ───────────
    print('Writing Fixed_Prerequisites.xlsx…')
    wb_pre = openpyxl.Workbook()

    # Sheet 1: Item Groups
    ws_ig = wb_pre.active
    ws_ig.title = 'Item Groups'
    ws_ig.append(['Item Group Name', 'Parent Item Group', 'Is Group'])
    style_header(ws_ig)
    for i, grp in enumerate(ITEM_GROUPS, 2):
        ws_ig.append([grp, 'All Item Groups', 0])
        if i % 2 == 0:
            for cell in ws_ig[i]: cell.fill = ALT_FILL
    set_col_widths(ws_ig, {'A': 28, 'B': 22, 'C': 10})

    # Sheet 2: UOM
    ws_uom = wb_pre.create_sheet('UOM')
    ws_uom.append(['UOM Name', 'Description'])
    style_header(ws_uom)
    uom_rows = [
        ('Bag',   'Bag / Bags'),
        ('Brass', 'Brass (100 cubic feet) — construction volume unit'),
        ('Pkt',   'Packet'),
        ('RFT',   'Running Feet'),
        ('Roll',  'Roll'),
    ]
    for i, (name, desc) in enumerate(uom_rows, 2):
        ws_uom.append([name, desc])
        if i % 2 == 0:
            for cell in ws_uom[i]: cell.fill = ALT_FILL
    set_col_widths(ws_uom, {'A': 16, 'B': 44})

    # Collect HSN codes from the parent source file
    hsn_set = set()
    for row in parent_rows[1:]:
        d = dict(zip(p_headers, row))
        hsn = str(d.get('HSN/SAC') or '').strip()
        if hsn and hsn != '0':
            hsn_set.add(hsn)

    # Sheet 3: GST HSN Codes
    ws_hsn = wb_pre.create_sheet('GST HSN Codes')
    ws_hsn.append(['HSN Code', 'Description'])
    style_header(ws_hsn)
    # Common descriptions for known codes
    HSN_DESC = {
        '25051000': 'Silica sands and quartz sands',
        '25232100': 'White Portland cement',
        '25232900': 'Portland cement (other)',
        '27101941': 'Lubricating oils',
        '32089099': 'Paints/varnishes — other',
        '38140000': 'Organic composite solvents / thinners',
        '39079100': 'Unsaturated polyesters',
        '39172390': 'Rigid tubes/pipes of plastics',
        '44121000': 'Plywood of bamboo',
        '44219099': 'Wood articles NEC',
        '48201000': 'Registers / account books / stationery',
        '62114300': 'Garments of man-made fibres',
        '68109900': 'Articles of cement/concrete — other',
        '69079000': 'Ceramic flags/paving/tiles — other',
        '72139900': 'Iron/steel bars and rods (hot-rolled) — other',
        '72179090': 'Wire of iron/steel — other',
        '73089090': 'Structures and parts — other',
        '73269099': 'Other articles of iron/steel — other',
        '84679900': 'Parts of tools for working in hand',
        '84714900': 'ADP machines (other than portable)',
        '84833000': 'Bearing housings / plain shaft bearings',
        '85364900': 'Relays for voltage > 60V',
        '90318000': 'Measuring/checking instruments NEC',
        '94036000': 'Furniture of other materials',
    }
    for i, hsn in enumerate(sorted(hsn_set), 2):
        ws_hsn.append([hsn, HSN_DESC.get(hsn, '')])
        if i % 2 == 0:
            for cell in ws_hsn[i]: cell.fill = ALT_FILL
    set_col_widths(ws_hsn, {'A': 16, 'B': 54})

    # Sheet 4: Brands (collect from parsed variant data)
    brand_set = set()
    for d, attrs, status in variant_parsed:
        b = (d.get('Brand') or '').strip()
        if b:
            brand_set.add(b)
    ws_brand = wb_pre.create_sheet('Brands')
    ws_brand.append(['Brand'])
    style_header(ws_brand)
    for i, brand in enumerate(sorted(brand_set), 2):
        ws_brand.append([brand])
        if i % 2 == 0:
            for cell in ws_brand[i]: cell.fill = ALT_FILL
    set_col_widths(ws_brand, {'A': 28})

    wb_pre.save(os.path.join(BASE_DIR, 'Fixed_Prerequisites.xlsx'))

    # ── Write File 2: Fixed Parent Material ───────────────────────────────────
    print('Writing Fixed_Parent_Material.xlsx…')
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = 'Items'

    out_headers_p = [
        'Item Code', 'Maintain Stock', 'Is Fixed Asset', 'Asset Category',
        'Description', 'End of Life', 'Item Group', 'Default Unit of Measure',
        'Shelf Life In Days', 'Safety Stock', 'Lead Time in days', 'HSN/SAC',
        'Has Variants', 'Attribute (Variant Attributes)',
    ]
    ws2.append(out_headers_p)
    style_header(ws2)

    row_num = 2
    for row in parent_rows[1:]:
        d = dict(zip(p_headers, row))
        item_code = (d.get('Item Code') or '').strip()
        if not item_code:
            continue

        # Attributes derived from variants; fall back to generic 'Size'
        attrs_for_parent = parent_attrs.get(item_code)
        if not attrs_for_parent:
            # Parent has no variants parsed → emit single placeholder row
            attrs_for_parent = ['Size']  # default; user should review

        raw_uom = (d.get('Default Unit of Measure') or '').strip()
        uom = UOM_NORMALIZE.get(raw_uom.upper(), raw_uom)

        first = True
        for attr_name in attrs_for_parent:
            if first:
                ws2.append([
                    d.get('Item Code'), d.get('Maintain Stock'), d.get('Is Fixed Asset'),
                    d.get('Asset Category'), d.get('Description'), d.get('End of Life'),
                    d.get('Item Group'), uom,
                    d.get('Shelf Life In Days'), d.get('Safety Stock'),
                    d.get('Lead Time in days'), d.get('HSN/SAC'),
                    d.get('Has Variants'), attr_name,
                ])
                first = False
            else:
                # Continuation rows: only attribute column is needed
                ws2.append(['', '', '', '', '', '', '', '', '', '', '', '', '', attr_name])
            if row_num % 2 == 0:
                for cell in ws2[row_num]:
                    cell.fill = ALT_FILL
            row_num += 1

    set_col_widths(ws2, {
        'A': 40, 'B': 15, 'C': 14, 'D': 18, 'E': 40, 'F': 12,
        'G': 22, 'H': 24, 'I': 18, 'J': 14, 'K': 18, 'L': 14,
        'M': 14, 'N': 30,
    })
    freeze_and_filter(ws2)
    wb2.save(os.path.join(BASE_DIR, 'Fixed_Parent_Material.xlsx'))

    # ── Write File 3: Fixed Variants ──────────────────────────────────────────
    print('Writing Fixed_Variants.xlsx…')
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active
    ws3.title = 'Variants'

    out_headers_v = [
        'Item Code', 'Item Group', 'Default Unit of Measure', 'Has Variants',
        'Maintain Stock', 'Is Fixed Asset', 'Asset Category', 'Brand',
        'Description', 'End of Life', 'Shelf Life In Days', 'Variant Of',
        'Variant Based On', 'Safety Stock', 'Lead Time in days', 'HSN/SAC',
        'Attribute (Variant Attributes)', 'Attribute Value (Variant Attributes)',
    ]
    ws3.append(out_headers_v)
    style_header(ws3)

    row_num = 2
    ok_count = review_count = no_variant_count = 0

    for d, attrs, status in variant_parsed:
        item_code  = (d.get('Item Code') or '').strip()
        variant_of = (d.get('Variant Of') or '').strip()

        if status == 'NO_VARIANT_OF':
            no_variant_count += 1
            raw_uom_s = (d.get('Default Unit of Measure') or '').strip()
            uom_s = UOM_NORMALIZE.get(raw_uom_s.upper(), raw_uom_s)
            ws3.append([
                d.get('Item Code'), d.get('Item Group'), uom_s,
                d.get('Has Variants'), d.get('Maintain Stock'), d.get('Is Fixed Asset'),
                d.get('Asset Category'), d.get('Brand'), d.get('Description'),
                d.get('End of Life'), d.get('Shelf Life In Days'), '',
                '', d.get('Safety Stock'), d.get('Lead Time in days'), d.get('HSN/SAC'),
                '', '',
            ])
            if row_num % 2 == 0:
                for cell in ws3[row_num]:
                    cell.fill = ALT_FILL
            row_num += 1
            continue

        if status == 'OK':
            ok_count += 1
        else:
            review_count += 1

        raw_uom_v = (d.get('Default Unit of Measure') or '').strip()
        uom_v = UOM_NORMALIZE.get(raw_uom_v.upper(), raw_uom_v)

        first = True
        for attr_name, attr_value in attrs:
            fill = WARN_FILL if status == 'REVIEW' else (ALT_FILL if row_num % 2 == 0 else None)

            if first:
                ws3.append([
                    d.get('Item Code'), d.get('Item Group'), uom_v,
                    d.get('Has Variants'), d.get('Maintain Stock'), d.get('Is Fixed Asset'),
                    d.get('Asset Category'), d.get('Brand'), d.get('Description'),
                    d.get('End of Life'), d.get('Shelf Life In Days'), variant_of,
                    'Item Attribute', d.get('Safety Stock'), d.get('Lead Time in days'),
                    d.get('HSN/SAC'), attr_name, attr_value,
                ])
                first = False
            else:
                # Continuation rows for additional attributes of same variant
                ws3.append([
                    '', '', '', '', '', '', '', '', '', '', '',
                    '', '', '', '', '', attr_name, attr_value,
                ])

            if fill:
                for cell in ws3[row_num]:
                    cell.fill = fill
            row_num += 1

    set_col_widths(ws3, {
        'A': 42, 'B': 22, 'C': 24, 'D': 14, 'E': 15, 'F': 14,
        'G': 18, 'H': 16, 'I': 40, 'J': 12, 'K': 18, 'L': 38,
        'M': 18, 'N': 14, 'O': 18, 'P': 14, 'Q': 32, 'R': 38,
    })
    freeze_and_filter(ws3)
    wb3.save(os.path.join(BASE_DIR, 'Fixed_Variants.xlsx'))

    # ── Final report ──────────────────────────────────────────────────────────
    total_with_parent = ok_count + review_count
    print('\n' + '=' * 60)
    print('DONE')
    print('=' * 60)
    print(f'  Unique attributes created : {len(attr_registry)}')
    print(f'  Total attribute values    : {sum(len(v) for v in attr_registry.values())}')
    print()
    print(f'  Variants parsed – OK      : {ok_count}')
    print(f'  Variants needing REVIEW   : {review_count}  (yellow in Fixed_Variants.xlsx)')
    print(f'  Items with no Variant Of  : {no_variant_count}  (standalone/orphan)')
    print()
    print(f'  Parse accuracy            : {ok_count/max(total_with_parent,1)*100:.1f}%')
    print()
    print('Output files:')
    print(f'  {BASE_DIR}/Fixed_Prerequisites.xlsx          ← IMPORT FIRST')
    print(f'  {BASE_DIR}/Fixed_Item_Attribute_Masters.xlsx ← IMPORT SECOND')
    print(f'  {BASE_DIR}/Fixed_Parent_Material.xlsx        ← IMPORT THIRD')
    print(f'  {BASE_DIR}/Fixed_Variants.xlsx               ← IMPORT LAST')

    print('\nTop REVIEW items (first 20):')
    shown = 0
    for d, attrs, status in variant_parsed:
        if status == 'REVIEW' and shown < 20:
            print(f"  [{d.get('Item Group')}] {d.get('Item Code')} → {attrs}")
            shown += 1


if __name__ == '__main__':
    main()
